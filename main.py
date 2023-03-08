import boto3
import jmespath
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook

class Roles_entities:

    def __init__(self, session):
        self.session = session
        self.client = session.client('ec2', region_name='eu-west-2')
        self.regions = self.client.describe_regions()
        self.r_list = []
    
    def return_df(self):
        self.describe_instances()
        r_json = json.dumps(self.r_list)
        r_df = pd.read_json(r_json)
        print("Roles dataframe complete")
        if r_df.empty:
            print("Roles dataframe empty")
        r_df = self.clean_lists(r_df)
        return r_df

    ## Clean list columns
    def clean_lists(self, r_df):
        r_df['TrustedEntity'] = r_df['TrustedEntity'].str.get(0)
        return r_df

    ## Describe instances
    def describe_instances(self):
        client = self.session.client('iam', region_name='eu-west-2')
        role_paginator = client.get_paginator('list_roles')
        self.filter_response(role_paginator)
    
    ## Filter roles
    def filter_response(self, role_paginator):
        for page in role_paginator.paginate():
            for role in page["Roles"]:
                role_name = jmespath.search("RoleName", role)
                trusted_entity = self.get_trusted_entity(role_name)
                role_dict = {
                    'RoleName':role_name,
                    'TrustedEntity':trusted_entity
                    }
                if role_dict:
                    self.r_list.append(role_dict)
                else:
                    pass

    ## Get trusted entity
    def get_trusted_entity(self, role_name):
        client = self.session.client('iam', region_name='eu-west-2')
        response = client.get_role(RoleName=role_name)
        trusted_entity = jmespath.search("Role.AssumeRolePolicyDocument.Statement[*].Principal.AWS", response)
        return trusted_entity

    ## Write to excel
    def make_blank_excel(self, profile):
        wb = openpyxl.Workbook()
        wb.save(profile+".xlsx")


    def write_new_sheet(self, profile, df, sheet_name):
        if df.empty:
            pass
        else:
            workbook = load_workbook(profile+".xlsx")
            writer = pd.ExcelWriter(profile+".xlsx")
            writer.book = workbook
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

# Run class test
def test_run():
    for profile in boto3.session.Session().available_profiles:
        if profile == "default":
            pass
        else:
            session = boto3.session.Session(profile_name=profile)
            test_run = Roles_entities(session)
            df = test_run.return_df()
            test_run.make_blank_excel(profile)
            test_run.write_new_sheet(profile, df, "Roles")
    
test_run()
