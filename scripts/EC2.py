import boto3
import jmespath
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook

class EC2:

    def __init__(self, session):
        self.session = session
        self.client = session.client('ec2', region_name='eu-west-1')
        self.regions = self.client.describe_regions()
        self.r_list = []
    
    def return_df(self):
        self.list_regions()
        r_json = json.dumps(self.r_list)
        r_df = pd.read_json(r_json)
        print("EC2 dataframe complete")
        if r_df.empty:
            print("EC2 dataframe empty")
        return r_df

    ## Loop through each region
    def list_regions(self):
        for region in self.regions['Regions']:
            region = region.get("RegionName")
            self.describe_instances(region)

    ## Describe instances
    def describe_instances(self, region):
        client = self.session.client('ec2', region_name=region)
        paginator = client.get_paginator('describe_instances')
        self.filter_response(paginator, region)

    # Filter respones
    def filter_response(self, paginator, region):
        for page in paginator.paginate():
            for r in page["Reservations"]:
                for i in r["Instances"]:
                    InstanceId = jmespath.search("InstanceId", i)
                    ssm_response = self.filter_for_ssm([InstanceId], region)
                    instance_dict = {
                        'Region':region,
                        'Instance ID':InstanceId,
                        'Name':(str(jmespath.search("Tags[?Key=='Name'].Value | [0]", i)).strip("[]")),
                        'Instance Type':jmespath.search("InstanceType", i),
                        'State':jmespath.search("State.Name", i),
                        'StateTransitionReason':jmespath.search("StateTransitionReason", i),
                        'Availability Zone':jmespath.search("Placement.AvailabilityZone", i),
                        'Private IP Address':jmespath.search("PrivateIpAddress", i),
                        'Public IP Address':jmespath.search("PublicIpAddress", i),
                        'Auto-Scaling-Group':(str(jmespath.search("Tags[?Key=='aws:autoscaling:groupName'].Value | [0]", i)).strip("[]")),
                        'Druva':(str(jmespath.search("Tags[?Key=='Backup'].Value | [0]", i)).strip("[]")),
                        'IsMonitored':(str(jmespath.search("Tags[?Key=='IsMonitored'].Value | [0]", i)).strip("[]")),
                        'Security Groups':(str(jmespath.search("SecurityGroups[*].GroupName | []", i)).strip("[]")).replace("'", ""),
                        'SSM Agent Version':ssm_response,
                        }
                    if instance_dict:
                        self.r_list.append(instance_dict)
                    else:
                        pass
    
    def filter_for_ssm(self, instances, region):
        client = self.session.client('ssm', region_name=region)
        response = client.describe_instance_information(
            Filters=[{'Key': 'InstanceIds', 'Values':instances}]
        )
        agent_version = str(jmespath.search("InstanceInformationList[*].AgentVersion", response)).strip("[]")
        return agent_version
    ## Write to excel
    def make_blank_excel(self, profile):
        wb = openpyxl.Workbook()
        wb.save(profile+".xlsx")


    def write_new_sheet(self, profile, df, sheet_name):
        if df.empty:
            pass
        else:
            workbook = load_workbook(profile+".xlsx")
            writer = pd.ExcelWriter(profile+".xlsx")
            writer.book = workbook
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

# Run class test
def test_run():
    for profile in boto3.session.Session().available_profiles:
        if profile == "default":
            pass
        else:
            session = boto3.session.Session(profile_name=profile)
            test_run = EC2(session)
            df = test_run.return_df()
            test_run.make_blank_excel(profile)
            test_run.write_new_sheet(profile, df, "EC2")
    
test_run()