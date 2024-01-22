import boto3
import jmespath
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook

class NetworkInterfaces:

    def __init__(self, session):
        self.session = session
        self.client = session.client('ec2', region_name='eu-west-2')
        self.regions = self.client.describe_regions()
        self.r_list = []
    
    def return_df(self):
        self.list_regions()
        r_json = json.dumps(self.r_list)
        r_df = pd.read_json(r_json)
        print("Network interface dataframe complete")
        if r_df.empty:
            print("Network interface dataframe empty")
        return r_df

    ## Loop through each region
    def list_regions(self):
        for region in self.regions['Regions']:
            region = region.get("RegionName")
            self.describe_instances(region)

    ## Describe instances
    def describe_instances(self, region):
        client = self.session.client('ec2', region_name=region)
        response = client.describe_network_interfaces()
        self.filter_response(response, region)

    # Filter respones
    def filter_response(self, response, region):
        for net_int in response:
            net_int_dict = {
                'Region':region,
                'AttachmentId':(str(jmespath.search("Attachment.AttachmentId", net_int)).strip("[]")),
                'AvailabilityZone':jmespath.search("AvailabilityZone", net_int),
                'VpcId':jmespath.search("VpcId", net_int),
                'SubnetId':jmespath.search("SubnetId", net_int),
                'Status':jmespath.search("Status", net_int),
                'PublicIp':(str(jmespath.search("Association.PublicIp", net_int)).strip("[]")),
                'PublicDnsName':(str(jmespath.search("Association.PublicDnsName", net_int)).strip("[]")),
                'Description':jmespath.search("Description", net_int),
                'NetworkInterfaceId':jmespath.search("NetworkInterfaceId", net_int),
                'OwnerId':jmespath.search("OwnerId", net_int),
                'NetworkInterfaceId':jmespath.search("NetworkInterfaceId", net_int),
                'PrivateDnsName':jmespath.search("PrivateDnsName", net_int),
                'PrivateIpAddress':jmespath.search("PrivateIpAddress", net_int),
                }
            if net_int_dict:
                self.r_list.append(net_int_dict)
            else:
                pass
## Write to excel
    def make_blank_excel(self, profile):
        wb = openpyxl.Workbook()
        wb.save(profile+".xlsx")


    def write_new_sheet(self, profile, df, sheet_name):
        if df.empty:
            pass
        else:
            workbook = load_workbook(profile+".xlsx")
            writer = pd.ExcelWriter(profile+".xlsx")
            writer.book = workbook
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

# Run class test
def test_run():
    for profile in boto3.session.Session().available_profiles:
        if profile == "default":
            pass
        else:
            session = boto3.session.Session(profile_name=profile)
            test_run = NetworkInterfaces(session)
            df = test_run.return_df()
            test_run.make_blank_excel(profile)
            test_run.write_new_sheet(profile, df, "NetworkInterfaces")
    
test_run()
