import boto3
import jmespath
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook


class Users:

    def __init__(self, session):
        self.session = session
        self.r_list = []
    
    def return_df(self):
        self.describe_instances()
        r_json = json.dumps(self.r_list)
        r_df = pd.read_json(r_json)
        print("User dataframe complete")
        if r_df.empty:
            print("User dataframe dataframe empty")
        return r_df

    ## Describe instances
    def describe_instances(self):
        client = self.session.client('iam')
        paginator = client.get_paginator('list_users')
        self.filter_response(paginator)

    # Filter respones
    def filter_response(self, paginator):
        for page in paginator.paginate():
            for user in page["Users"]:
                access_key = self.get_access_key_age(user["UserName"])
                if not access_key:
                    access_key = ["", "", "", "", ""]
                else:
                    user_dict = {
                        'UserName':jmespath.search("UserName", user),
                        'CreateDate':str(jmespath.search("CreateDate", user)),
                        'UserId':jmespath.search("UserId", user),
                        'Arn':jmespath.search("Arn", user),
                        'Key Status':access_key[0],
                        'Key Age':access_key[1],
                        'Key Last Used':access_key[2],
                        'Last Service Used':access_key[3],
                        'Last Region Used':access_key[4],
                        }
                    if user_dict:
                        self.r_list.append(user_dict)
                    else:
                        pass

    
    def get_access_key_age(self, username):
        client = self.session.client('iam')
        paginator = client.get_paginator('list_access_keys')
        for page in paginator.paginate(UserName=username, ):
            for user in page["AccessKeyMetadata"]:
                status = str(jmespath.search("Status", user))
                create_date = str(jmespath.search("CreateDate", user))
                key_id = self.key_last_used(jmespath.search("AccessKeyId", user))
                key_last_used = key_id[0]
                last_service = key_id[1]
                last_region = key_id[2]
                return status, create_date, key_last_used, last_service, last_region
    
    def key_last_used(self, key_id):
        client = self.session.client('iam')
        response = client.get_access_key_last_used(AccessKeyId=key_id)    
        key_last_used = str(jmespath.search("AccessKeyLastUsed.LastUsedDate", response))
        last_service = str(jmespath.search("AccessKeyLastUsed.ServiceName", response))
        last_region = str(jmespath.search("AccessKeyLastUsed.Region", response))
        return key_last_used, last_service, last_region
## Write to excel
    def make_blank_excel(self, profile):
        wb = openpyxl.Workbook()
        wb.save(profile+".xlsx")


    def write_new_sheet(self, profile, df, sheet_name):
        if df.empty:
            pass
        else:
            workbook = load_workbook(profile+".xlsx")
            writer = pd.ExcelWriter(profile+".xlsx")
            writer.book = workbook
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

# Run class test
def test_run():
    for profile in boto3.session.Session().available_profiles:
        if profile == "default":
            pass
        else:
            session = boto3.session.Session(profile_name=profile)
            test_run = Users(session)
            df = test_run.return_df()
            test_run.make_blank_excel(profile)
            test_run.write_new_sheet(profile, df, "Users")
    
test_run()