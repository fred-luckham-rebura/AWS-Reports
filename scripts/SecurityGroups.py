import boto3
import jmespath
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook

class Security_Groups:

    def __init__(self, session):
        self.session = session
        self.client = session.client('ec2', region_name='eu-west-1')
        self.regions = self.client.describe_regions()
        self.r_list = []
    
    def return_df(self):
        self.list_regions()
        r_json = json.dumps(self.r_list)
        r_df = pd.read_json(r_json)
        print("SecurityGroups dataframe complete")
        if r_df.empty:
            print("SecurityGroups dataframe empty")
        return r_df

    ## Loop through each region
    def list_regions(self):
        for region in self.regions['Regions']:
            region = region.get("RegionName")
            self.describe_instances(region)

    ## Describe instances
    def describe_instances(self, region):
        client = self.session.client('ec2', region_name=region)
        paginator = client.get_paginator('describe_security_groups')
        self.filter_response(paginator, region)

    # Filter respones
    def filter_response(self, paginator, region):
        for page in paginator.paginate():
            for sg in page["SecurityGroups"]:
                sg_dict = {
                    'Region':region,
                    'Description':jmespath.search("Description", sg),
                    }
                if sg_dict:
                    self.r_list.append(sg_dict)
                else:
                    pass
    ## Write to excel
    def make_blank_excel(self, profile):
        wb = openpyxl.Workbook()
        wb.save(profile+".xlsx")


    def write_new_sheet(self, profile, df, sheet_name):
        if df.empty:
            pass
        else:
            workbook = load_workbook(profile+".xlsx")
            writer = pd.ExcelWriter(profile+".xlsx")
            writer.book = workbook
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

# Run class test
def test_run():
    for profile in boto3.session.Session().available_profiles:
        if profile == "default":
            pass
        else:
            session = boto3.session.Session(profile_name=profile)
            test_run = Security_Groups(session)
            df = test_run.return_df()
            test_run.make_blank_excel(profile)
            test_run.write_new_sheet(profile, df, "SG")
    
test_run()